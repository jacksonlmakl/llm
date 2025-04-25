import os
import sys
import pandas as pd
from PIL import Image
import pytesseract
import docx2txt
import PyPDF2
import csv
import openpyxl
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import TextLoader
from langchain_community.docstore.document import Document
from transformers import pipeline
from tqdm import tqdm
import logging
import hashlib
import json
import time

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DocumentProcessor:
    def __init__(self, directory_path):
        logging.disable(logging.CRITICAL)  # Disable all logging
        self.directory_path = directory_path
        # Initialize the embedding model
        self.embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
        
        # Initialize text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            separators=["\n\n", "\n", " ", ""]
        )
        
        # Initialize document storage
        self.documents = []
        
        # Initialize Tesseract for OCR (make sure tesseract is installed)
        pytesseract.pytesseract.tesseract_cmd = r'tesseract'  # Update this path if needed
        
        # Track processed files for incremental updates
        self.processed_files = {}
        
    def load_file_registry(self, registry_path):
        """Load the registry of processed files"""
        if os.path.exists(registry_path):
            try:
                with open(registry_path, 'r') as f:
                    self.processed_files = json.load(f)
                logger.info(f"Loaded registry with {len(self.processed_files)} processed files")
                return True
            except Exception as e:
                logger.error(f"Error loading file registry: {str(e)}")
        return False
    
    def save_file_registry(self, registry_path):
        """Save the registry of processed files"""
        try:
            with open(registry_path, 'w') as f:
                json.dump(self.processed_files, f)
            logger.info(f"Saved registry with {len(self.processed_files)} processed files")
            return True
        except Exception as e:
            logger.error(f"Error saving file registry: {str(e)}")
            return False
            
    def get_file_hash(self, filepath):
        """Get hash of file to detect changes"""
        try:
            file_stat = os.stat(filepath)
            # Use modified time and size as a quick hash
            return f"{file_stat.st_mtime}_{file_stat.st_size}"
        except Exception as e:
            logger.error(f"Error getting file hash for {filepath}: {str(e)}")
            return None
    
    def process_directory(self, registry_path):
        """Process all files in the directory and extract text, skipping already processed files"""
        logger.info(f"Processing directory: {self.directory_path}")
        
        if not os.path.exists(self.directory_path):
            logger.error(f"Directory not found: {self.directory_path}")
            return False
        
        # Load registry of processed files
        self.load_file_registry(registry_path)
            
        file_count = 0
        new_files_count = 0
        updated_files_count = 0
        
        for filename in tqdm(os.listdir(self.directory_path)):
            filepath = os.path.join(self.directory_path, filename)
            if os.path.isfile(filepath):
                try:
                    # Get file hash to detect changes
                    file_hash = self.get_file_hash(filepath)
                    
                    # Skip if file hasn't changed
                    if filename in self.processed_files and self.processed_files[filename]["hash"] == file_hash:
                        # Add the already processed content to documents list
                        self.documents.append(Document(
                            page_content=self.processed_files[filename]["content"], 
                            metadata={"source": filename}
                        ))
                        file_count += 1
                        continue
                    
                    # Process the file if it's new or changed
                    file_extension = os.path.splitext(filename)[1].lower()
                    text = self.extract_text(filepath, file_extension)
                    
                    if text:
                        self.documents.append(Document(page_content=text, metadata={"source": filename}))
                        file_count += 1
                        
                        # Update the registry
                        if filename in self.processed_files:
                            updated_files_count += 1
                        else:
                            new_files_count += 1
                            
                        self.processed_files[filename] = {
                            "hash": file_hash,
                            "content": text,
                            "processed_time": time.time()
                        }
                except Exception as e:
                    logger.error(f"Error processing {filename}: {str(e)}")
        
        # Save the updated registry
        self.save_file_registry(registry_path)
        
        logger.info(f"Successfully processed {file_count} files in total")
        logger.info(f"New files: {new_files_count}, Updated files: {updated_files_count}")
        return True
    
    def extract_text(self, filepath, file_extension):
        """Extract text from different file types"""
        logger.debug(f"Extracting text from {filepath} with extension {file_extension}")
        
        if file_extension in ['.txt']:
            return self.extract_from_txt(filepath)
        elif file_extension in ['.pdf']:
            return self.extract_from_pdf(filepath)
        elif file_extension in ['.docx', '.doc']:
            return self.extract_from_docx(filepath)
        elif file_extension in ['.csv']:
            return self.extract_from_csv(filepath)
        elif file_extension in ['.xlsx', '.xls']:
            return self.extract_from_xlsx(filepath)
        elif file_extension in ['.jpg', '.jpeg', '.png']:
            return self.extract_from_image(filepath)
        else:
            logger.warning(f"Unsupported file type: {file_extension}")
            return None
    
    # All the extraction methods remain the same
    def extract_from_txt(self, filepath):
        # Same implementation as before
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error extracting text from {filepath}: {str(e)}")
            return ""
    
    def extract_from_pdf(self, filepath):
        # Same implementation as before
        try:
            text = ""
            with open(filepath, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF {filepath}: {str(e)}")
            return ""
    
    # Other extraction methods remain the same...
    def extract_from_docx(self, filepath):
        try:
            return docx2txt.process(filepath)
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {filepath}: {str(e)}")
            return ""
    
    def extract_from_csv(self, filepath):
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                csv_reader = csv.reader(file)
                rows = list(csv_reader)
                
                # Convert CSV to a readable text format
                headers = rows[0] if rows else []
                text = "CSV File Content:\n"
                
                # Add headers
                text += "Headers: " + ", ".join(headers) + "\n\n"
                
                # Add data rows
                for i, row in enumerate(rows[1:], start=1):
                    row_text = f"Row {i}: "
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            row_text += f"{headers[j]}: {cell}, "
                    text += row_text + "\n"
                    
                return text
        except Exception as e:
            logger.error(f"Error extracting text from CSV {filepath}: {str(e)}")
            return ""
    
    def extract_from_xlsx(self, filepath):
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            text = f"Excel File: {os.path.basename(filepath)}\n\n"
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                # Get all values from the sheet
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    text += "Empty sheet\n\n"
                    continue
                    
                # Extract headers (first row)
                headers = [str(cell) if cell is not None else "" for cell in data[0]]
                
                # Process data rows
                for i, row in enumerate(data[1:], start=2):
                    row_text = f"Row {i-1}: "
                    for j, cell in enumerate(row):
                        if j < len(headers) and headers[j]:
                            row_text += f"{headers[j]}: {str(cell) if cell is not None else 'N/A'}, "
                    text += row_text + "\n"
                
                text += "\n"
            
            return text
        except Exception as e:
            logger.error(f"Error extracting text from Excel {filepath}: {str(e)}")
            return ""
    
    def extract_from_image(self, filepath):
        try:
            img = Image.open(filepath)
            text = pytesseract.image_to_string(img)
            if not text:
                return f"Image file without extractable text: {os.path.basename(filepath)}"
            return text
        except Exception as e:
            logger.error(f"Error extracting text from image {filepath}: {str(e)}")
            return ""
    
    def create_vector_store(self, save_path=None, update_existing=False):
        """Create or update a vector store from chunked documents"""
        if not self.documents:
            logger.warning("No documents to process.")
            return None
            
        logger.info("Chunking documents...")
        chunked_docs = []
        
        for doc in tqdm(self.documents):
            chunks = self.text_splitter.split_text(doc.page_content)
            for i, chunk in enumerate(chunks):
                metadata = doc.metadata.copy()
                metadata["chunk"] = i
                chunked_docs.append(Document(page_content=chunk, metadata=metadata))
        
        logger.info(f"Created {len(chunked_docs)} chunks from {len(self.documents)} documents")
        
        if not chunked_docs:
            logger.warning("No chunks created.")
            return None
            
        # Create or update vector store
        if update_existing and save_path and os.path.exists(os.path.join(save_path, "index.faiss")):
            logger.info(f"Loading existing vector store from {save_path}")
            vector_store = FAISS.load_local(save_path, self.embeddings, allow_dangerous_deserialization=True)
            
            # Add new documents to existing vector store
            if chunked_docs:
                logger.info(f"Updating vector store with {len(chunked_docs)} new chunks")
                vector_store.add_documents(chunked_docs)
                # Save the updated vector store
                vector_store.save_local(save_path)
        else:
            logger.info("Creating new vector store...")
            vector_store = FAISS.from_documents(chunked_docs, self.embeddings)
            
            # Save the vector store if a path is provided
            if save_path:
                logger.info(f"Saving vector store to {save_path}")
                vector_store.save_local(save_path)
        
        return vector_store

class RAGSystem:
    def __init__(self, vector_store=None, vector_store_path=None):
        """Initialize the RAG system"""
        if vector_store:
            self.vector_store = vector_store
        elif vector_store_path and os.path.exists(os.path.join(vector_store_path, "index.faiss")):
            self.embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
            self.vector_store = FAISS.load_local(vector_store_path, self.embeddings, allow_dangerous_deserialization=True)
        else:
            raise ValueError("Either vector_store or a valid vector_store_path must be provided")
    
    def query(self, prompt, k=5):
        """
        Query the vector store and return the prompt enriched with relevant information
        
        Args:
            prompt: The user's prompt/question
            k: Number of relevant chunks to retrieve
            
        Returns:
            Enriched prompt with relevant context
        """
        # Retrieve relevant documents
        retrieved_docs = self.vector_store.similarity_search(prompt, k=k)
        
        # Extract relevant context
        context = "\n\n".join([doc.page_content for doc in retrieved_docs])
        
        # Create sources list for citation
        sources = []
        for doc in retrieved_docs:
            source = doc.metadata.get("source", "Unknown")
            if source not in sources:
                sources.append(source)
        
        # Format sources
        sources_text = "\n".join([f"- {source}" for source in sources])
        
        # Combine context with the prompt for a better response
        enriched_prompt = f"""
        Question/Prompt: {prompt}
        
        Here is relevant information that may help answer the question:
        
        {context}
        
        Sources consulted:
        {sources_text}
        
        Based on the above information, here's a comprehensive answer to the original question/prompt:
        """
                
        return enriched_prompt

        
def RAG(prompt, directory_path="documents", vector_store_path="vector_store", file_registry_path="file_registry.json"):
    """
    Retrieval-Augmented Generation function that only processes new or changed files
    
    Args:
        prompt: The user's prompt/question
        directory_path: Path to the directory containing documents
        vector_store_path: Path to save/load the vector store
        file_registry_path: Path to save/load the file registry
        
    Returns:
        Enriched prompt with relevant context
    """
    # Check if vector store exists
    vector_store_exists = os.path.exists(os.path.join(vector_store_path, "index.faiss"))
    
    # Process documents and create/update vector store
    processor = DocumentProcessor(directory_path)
    
    # Process the directory, skipping unchanged files
    processor.process_directory(file_registry_path)
    
    # Create or update vector store
    vector_store = processor.create_vector_store(
        save_path=vector_store_path,
        update_existing=vector_store_exists
    )
    
    # Create RAG system
    rag_system = RAGSystem(vector_store=vector_store)
    
    # Query the RAG system
    enriched_prompt = rag_system.query(prompt).replace("\n","").strip()
    return enriched_prompt